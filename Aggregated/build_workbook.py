"""
Turn the flat player/team CSVs that build_season_aggregate.py produces into
one .xlsx workbook, split across category tabs (Passing, Progression,
Creativity, Duels & Defensive Actions, Shooting, Set Pieces, Discipline,
Goalkeeping, Splits, the two "New Metrics" tabs, Team Season, and a Glossary
tab) using the same category rules in column_layout.py -- so the tab split
can never drift from what the build script actually produced.

Usage: python3 build_workbook.py [season]   (default: 2025-2026; requires
       build_season_aggregate.py to have already been run for that season)
"""
import csv
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from column_layout import categorize, PLAYER_TAB_RULES, TEAM_TAB_RULES
from display_names import display_name

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SEASON = sys.argv[1] if len(sys.argv) > 1 else "2025-2026"
OUT_DIR = os.path.join(ROOT, "Aggregated", SEASON)

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
IDENTITY_FILL = PatternFill("solid", fgColor="EEF2F7")
MAX_SHEET_NAME = 31  # Excel hard limit

GLOSSARY = """Eredivisie {season} -- aggregated player & team statistics
Built by Aggregated/build_season_aggregate.py + build_workbook.py

WHAT THIS IS
One Wyscout-style stat sheet per player and per team for the season, split
across tabs by category. Two different origins feed the same row:
  - Wyscout-style counting stats (passing, crossing, duels, defensive
    actions, discipline, goalkeeping, shooting, set pieces, key passes/
    assists/xA, progression, splits) are computed directly from the raw
    Opta event stream in Events/{season}/*.json.
  - The two "New Metrics" tabs are joined in from this repo's existing
    model outputs (xT/, GDA/, Disruption/CSV/, Box Entry Models/,
    Cross Models/), not recomputed here.

Row 1 on every tab is a concise display name (e.g. "PAdj Tackles /90");
row 2 (small, grey) is the exact snake_case CSV column it maps to, in case
a formula needs the unambiguous name. See display_names.py for the mapping.

DEFINITIONS WORTH KNOWING BEFORE YOU USE A COLUMN
- Progressive pass/carry: cuts the distance to the opponent's goal by at
  least 30 yards if both ends are in the player's own half, 15 yards if it
  crosses into the attacking half, or 10 yards if both ends are already in
  the attacking half. A common public heuristic, not unique to this repo.
  KNOWN ISSUE: goalkeepers can top progressive-pass leaderboards because
  long defensive distribution satisfies this formula the same way a
  genuine line-breaking pass does. Filter to outfield players before using
  this for scouting.
- Carries: NOT tracking data. Inferred by chaining a player's own
  consecutive ball-touching events (pass/take-on/shot/ball-touch) when the
  gap is <=8 seconds and the ball moved >=3m, and the chain resets after a
  dead-ball restart (corner/free-kick/throw-in) or a failed pass. This is a
  heuristic approximation, the same kind of inference xT/xt_model_meta.json
  already relies on for its own carry detection -- expect it to undercount
  or overcount in individual matches.
- Key pass / Assist / xA: for every shot, walk back up to 4 events for the
  nearest completed same-team pass. xA sums the shot's own xG onto the
  passer regardless of outcome; Assists only count shots that scored. A bug
  in the original version (locating each shot via the Danger CSV's own
  event_index, which turns out not to match its true position in the event
  list) was found and fixed while making this pipeline season-portable --
  see README.md for the details.
- By delivery type (Cutback/Cross/Through Ball/Set Piece/Open Play on the
  Creativity tab): the same qualifying pass is classified by what it was
  (pull-back, cross, through ball, free-kick/corner, else open play), so
  "who creates from cutbacks" and "who creates from open play" don't blend
  into one number. The five buckets sum exactly to the unbroken total.
- Shot-Creating Actions (SCA) / Goal-Creating Actions (GCA): up to the 2
  most recent successful actions (completed pass, successful take-on, or
  a foul won) by the shooting team before a shot/goal, credited to up to 2
  different players. Modelled after the FBref definition, computed here
  from this feed's own events -- not sourced from FBref.
- Passes received: approximated as the very next on-ball action by a
  teammate within 5 seconds of a completed pass. This feed has no explicit
  "intended receiver" tag, so misplaced/deflected passes or a receiver who
  needed a 6th second will be missed -- an approximation, not a certainty.
- Crosses (raw, "crosses_attempted" etc.) count every pass carrying the
  cross qualifier, including set pieces. This is a DIFFERENT universe from
  the "xp_cross_*" columns (New Metrics tab), which come from Cross
  Models/xp_player_leaderboard_eredivisie.csv and may be scoped to
  open-play crosses only -- the two will not match, by design, and that is
  not a data error.
- PAdj (possession-adjusted) tackles/interceptions/clearances/aerial
  duels/ball recoveries: raw per-90 rate multiplied by
  (league-average opponent-possession% / this team's own opponent-
  possession%), where team possession% is this team's average share of
  total match pass attempts across its own matches. Teams that see less of
  the ball face more raw defensive opportunities, so PAdj scales a
  low-possession team's defender up and a high-possession team's defender
  down for a fairer read across different team styles.
- goal_difference_added (GDA) mixes on-pitch goal differential with an
  action-value model (see GDA/gda_model_meta.json) -- a player on a poor
  team can show a large negative value even with good individual
  performances, because it partly reflects team goals conceded while they
  were on the pitch.

WHAT'S DELIBERATELY NOT IN HERE
No composite score. Every component here is per-90 (and the core "new
metrics" are z-scored among reliable_sample players) -- enough to build a
weighted composite on top, but combining them into one number is a
modelling choice (which weights, validated how) that hasn't been through
face validity (video review), construct validity (does it just track
possession share / position / minutes?), or weight-sensitivity testing.
No shrinkage / confidence intervals -- reliable_sample (minutes >= 450,
~5 full matches) is a blunt cutoff, not empirical-Bayes shrinkage.
No cross-season or predictive-validity check -- this is one season's
descriptive aggregate.

Team `matches` differ (34 vs 35 vs 36) because of the Eredivisie's
end-of-season European-qualification play-offs, not a data error.
"""


def read_csv(path):
    with open(path, encoding="utf-8-sig", newline="") as f:
        r = csv.reader(f)
        header = next(r)
        rows = list(r)
    return header, rows


def autosize(ws, max_width=28):
    for i, col in enumerate(ws.columns, start=1):
        width = min(max_width, max((len(str(c.value)) if c.value is not None else 0) for c in col) + 2)
        ws.column_dimensions[get_column_letter(i)].width = max(width, 8)


RAW_NAME_FONT = Font(color="9AA5B1", italic=True, size=8)


def write_sheet(wb, title, raw_headers, rows, identity_count):
    """Row 1 = concise display name (bold). Row 2 = the exact snake_case CSV
    column name (small, grey) so a formula or a question about "which raw
    field is this" always has an answer without leaving the tab."""
    title = title[:MAX_SHEET_NAME]
    ws = wb.create_sheet(title)
    ws.append([display_name(c) for c in raw_headers])
    ws.append(raw_headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for cell in ws[2]:
        cell.font = RAW_NAME_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    if identity_count:
        for r in range(1, ws.max_row + 1):
            for c in range(1, identity_count + 1):
                ws.cell(row=r, column=c).fill = IDENTITY_FILL if r > 2 else HEADER_FILL
    ws.freeze_panes = get_column_letter(identity_count + 1) + "3"
    ws.row_dimensions[1].height = 30
    autosize(ws)
    return ws


def build_tabbed_sheets(wb, csv_path, rules, identity_cols, tab_prefix=""):
    header, rows = read_csv(csv_path)
    col_index = {name: i for i, name in enumerate(header)}
    tabs = categorize(header, rules, identity_cols=identity_cols)
    id_idx = [col_index[c] for c in identity_cols]
    for tab_name, cols in tabs:
        idx = [col_index[c] for c in cols]
        sheet_headers = identity_cols + cols
        sheet_rows = [[row[i] for i in id_idx] + [row[i] for i in idx] for row in rows]
        write_sheet(wb, f"{tab_prefix}{tab_name}", sheet_headers, sheet_rows, len(identity_cols))
    return [name for name, _ in tabs]


def main():
    player_csv = os.path.join(OUT_DIR, "player_season_aggregated.csv")
    team_csv = os.path.join(OUT_DIR, "team_season_aggregated.csv")
    out_path = os.path.join(OUT_DIR, f"eredivisie_{SEASON}_aggregated.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    player_tabs = build_tabbed_sheets(wb, player_csv, PLAYER_TAB_RULES,
                                       identity_cols=["player_id", "player_name", "team_name"],
                                       tab_prefix="P - ")
    team_tabs = build_tabbed_sheets(wb, team_csv, TEAM_TAB_RULES,
                                     identity_cols=["team_name"], tab_prefix="T - ")

    glossary = wb.create_sheet("Glossary", 0)
    for line in GLOSSARY.format(season=SEASON).split("\n"):
        glossary.append([line])
    glossary.column_dimensions["A"].width = 100
    for row in glossary.iter_rows():
        row[0].alignment = Alignment(wrap_text=False)
    glossary["A1"].font = Font(bold=True, size=14)

    wb.save(out_path)
    total_metric_cols = len(read_csv(player_csv)[0]) - 3
    print(f"Wrote {out_path}")
    print(f"  Player tabs: {player_tabs}")
    print(f"  Team tabs:   {team_tabs}")
    print(f"  {total_metric_cols} player metric columns across {len(player_tabs)} tabs")


if __name__ == "__main__":
    main()
