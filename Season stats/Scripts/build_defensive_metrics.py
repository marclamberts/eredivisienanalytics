"""Build Season stats/Defensive Metrics.xlsx from the season's event-derived
defensive actions (Events/*.json), plus season minutes (GDA/gda_player_summary.csv)
and position (same lineup-qualifier method as build_shooting_metrics.py /
build_passing_metrics.py). Two tabs: Per 90 and Total.

Defensive-action type IDs (tackle/interception/clearance/aerial/recovery/
blocked pass) match the DEF_TYPES set already used in PSV Season Report/
Scripts/season_expansion_pitches.py; the own-box threshold (x<=17) matches
that script's box_defending zone.
"""
import csv
import glob
import json
import os
from collections import defaultdict, Counter

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUT_DIR = os.path.join(ROOT, "Season stats")
OUT_PATH = os.path.join(OUT_DIR, "Defensive Metrics.xlsx")

DEF_THIRD_END, FINAL_THIRD_START = 100 / 3, 200 / 3
OWN_BOX_X, OWN_BOX_Y_LO, OWN_BOX_Y_HI = 17.0, 21.0, 79.0  # season_expansion_pitches.py box_defending

TACKLE, INTERCEPTION, CLEARANCE, FOUL, CARD, AERIAL = 7, 8, 12, 4, 17, 44
RECOVERY, DISPOSSESSED, ERROR, OFFSIDE_PROVOKED, BLOCKED_PASS = 49, 50, 51, 55, 74
DEFENSIVE_ACTION_TYPES = {TACKLE, INTERCEPTION, CLEARANCE, AERIAL, RECOVERY, BLOCKED_PASS}

LINEUP_TYPE_ID = 34
POS_MAP = {1: "GK", 2: "DEF", 3: "MID", 4: "FWD"}
TOUCH_EXCLUDE = {18, 19, 30, 32, 34, 37, 40, 70, 71, 90, 91}  # non-positional event types


def rows(path):
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def qmap(e):
    return {q.get("qualifierId"): q.get("value") for q in e.get("qualifier", []) or []}


def origin_third(x):
    if x < DEF_THIRD_END:
        return "Defensive Third"
    if x < FINAL_THIRD_START:
        return "Middle Third"
    return "Final Third"


def in_own_box(x, y):
    return x <= OWN_BOX_X and OWN_BOX_Y_LO <= y <= OWN_BOX_Y_HI


def specific_position_labels(n, group):
    """Left-to-right role names for `n` starters ranked by average touch
    width within `group`. Approximate (there is no official per-role tag in
    the data) -- see build_positions() docstring."""
    if group == "DEF":
        table = {1: ["CB"], 2: ["LB", "RB"], 3: ["LCB", "CB3", "RCB"],
                 4: ["LB", "LCB", "RCB", "RB"], 5: ["LWB", "LCB", "CB", "RCB", "RWB"]}
    elif group == "MID":
        table = {1: ["CM"], 2: ["LCM", "RCM"], 3: ["LM", "CM", "RM"],
                 4: ["LM", "LCM", "RCM", "RM"], 5: ["LM", "LCM", "CM", "RCM", "RM"]}
    elif group == "FWD":
        table = {1: ["ST"], 2: ["LST", "RST"], 3: ["LW", "ST", "RW"]}
    else:
        table = {}
    return table.get(n, [f"{group}{i + 1}" for i in range(n)])


def build_positions():
    """player_id -> (position group, specific position). See the identical
    helper in build_shooting_metrics.py / build_passing_metrics.py for the
    full derivation notes (qualifier 44/131/30 on the Team Set Up event,
    specific role approximated from average touch width per match)."""
    votes_group = defaultdict(Counter)
    votes_specific = defaultdict(Counter)
    for path in sorted(glob.glob(os.path.join(ROOT, "Events", "*.json"))):
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        events = data.get("event", [])

        touch_y = defaultdict(list)
        for e in events:
            pid, tid, y = e.get("playerId"), e.get("typeId"), e.get("y")
            if pid and y is not None and tid not in TOUCH_EXCLUDE:
                touch_y[pid].append(float(y))

        for e in events:
            if e.get("typeId") != LINEUP_TYPE_ID:
                continue
            q = qmap(e)
            pids = (q.get(30) or "").split(", ")
            pos = (q.get(44) or "").split(", ")
            slots = (q.get(131) or "").split(", ")
            if not (len(pids) == len(pos) == len(slots)):
                continue

            by_group = defaultdict(list)
            for pid, p, s in zip(pids, pos, slots):
                if s.strip() == "0" or p.strip() not in ("1", "2", "3", "4"):
                    continue
                p = int(p)
                votes_group[pid][p] += 1
                if p == 1:
                    votes_specific[pid]["GK"] += 1
                    continue
                avg_y = sum(touch_y[pid]) / len(touch_y[pid]) if touch_y.get(pid) else 50.0
                by_group[POS_MAP[p]].append((pid, avg_y))

            for group, players in by_group.items():
                players.sort(key=lambda t: t[1])
                labels = specific_position_labels(len(players), group)
                for (pid, _), label in zip(players, labels):
                    votes_specific[pid][label] += 1

    groups = {pid: POS_MAP[c.most_common(1)[0][0]] for pid, c in votes_group.items()}
    specific = {pid: c.most_common(1)[0][0] for pid, c in votes_specific.items()}
    return groups, specific


# --- team names -------------------------------------------------------------
team_names = {}
for r in rows(os.path.join(ROOT, "xT", "xt_team_summary.csv")):
    team_names[r["contestant_id"]] = r["team_name"]

# --- minutes / appearances per player (season totals) -----------------------
minutes = defaultdict(lambda: {"name": "", "contestant_ids": Counter(), "minutes": 0.0, "matches": 0})
for r in rows(os.path.join(ROOT, "GDA", "gda_player_summary.csv")):
    pid = r["player_id"]
    m = minutes[pid]
    m["name"] = r["player_name"]
    m["contestant_ids"][r["contestant_id"]] += 1
    m["minutes"] += num(r["minutes"])
    m["matches"] += int(num(r["matches"]))

position_groups, positions = build_positions()

ZONES = ["Defensive Third", "Middle Third", "Final Third"]


def zero_split(keys):
    return {k: 0 for k in keys}


agg = defaultdict(lambda: {
    "name": "", "contestant_ids": Counter(),
    "tackles": 0, "tackles_won": 0, "dribbled_past": 0,
    "interceptions": 0, "clearances": 0,
    "aerials": 0, "aerials_won": 0,
    "recoveries": 0, "blocked_passes": 0,
    "fouls_committed": 0, "fouls_won": 0, "cards": 0,
    "offsides_won": 0, "errors": 0,
    "def_actions": 0, "pressures": 0, "own_box_actions": 0,
    "zone": zero_split(ZONES),
})

# --- single pass over every match's events -----------------------------------
for path in sorted(glob.glob(os.path.join(ROOT, "Events", "*.json"))):
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    for e in data.get("event", []):
        pid = e.get("playerId")
        tid = e.get("typeId")
        if not pid or tid not in (DEFENSIVE_ACTION_TYPES | {FOUL, CARD, ERROR, OFFSIDE_PROVOKED}):
            continue
        a = agg[pid]
        a["name"] = e.get("playerName") or a["name"] or "Unknown"
        a["contestant_ids"][e.get("contestantId")] += 1

        x = num(e.get("x"))
        y = num(e.get("y"))
        success = e.get("outcome") == 1

        if tid == TACKLE:
            a["tackles"] += 1
            if success:
                a["tackles_won"] += 1
            else:
                a["dribbled_past"] += 1
        elif tid == INTERCEPTION:
            a["interceptions"] += 1
        elif tid == CLEARANCE:
            a["clearances"] += 1
        elif tid == AERIAL:
            a["aerials"] += 1
            if success:
                a["aerials_won"] += 1
        elif tid == RECOVERY:
            a["recoveries"] += 1
        elif tid == BLOCKED_PASS:
            a["blocked_passes"] += 1
        elif tid == FOUL:
            if success:
                a["fouls_committed"] += 1
            else:
                a["fouls_won"] += 1
        elif tid == CARD:
            a["cards"] += 1
        elif tid == ERROR:
            a["errors"] += 1
        elif tid == OFFSIDE_PROVOKED:
            a["offsides_won"] += 1

        if tid in DEFENSIVE_ACTION_TYPES:
            a["def_actions"] += 1
            a["zone"][origin_third(x)] += 1
            if x >= 50:
                a["pressures"] += 1
            if in_own_box(x, y):
                a["own_box_actions"] += 1

# --- player identity (name / team / matches / minutes), shared by every tab -
player_info = {}
for pid, a in agg.items():
    m = minutes.get(pid)
    cid_counter = m["contestant_ids"] if m and m["contestant_ids"] else a["contestant_ids"]
    contestant_id = cid_counter.most_common(1)[0][0] if cid_counter else None
    player_info[pid] = {
        "name": (m["name"] if m else "") or a["name"],
        "team": team_names.get(contestant_id, "Unknown"),
        "position_group": position_groups.get(pid, "Unknown"),
        "position": positions.get(pid, "Unknown"),
        "matches": m["matches"] if m else 0,
        "minutes": m["minutes"] if m else 0.0,
    }

# --- assemble one row per player ---------------------------------------------
records = []
for pid, a in agg.items():
    info = player_info[pid]
    name, team, position_group, position, matches, mins = (
        info["name"], info["team"], info["position_group"], info["position"], info["matches"], info["minutes"]
    )

    record = {
        "Player": name,
        "Team": team,
        "Position Group": position_group,
        "Position": position,
        "Matches": matches,
        "Minutes": round(mins, 1),
        "Defensive Actions": a["def_actions"],
        "Tackles": a["tackles"],
        "Tackles Won": a["tackles_won"],
        "Tackle Win %": round(a["tackles_won"] / a["tackles"] * 100, 1) if a["tackles"] else 0.0,
        "Dribbled Past": a["dribbled_past"],
        "Interceptions": a["interceptions"],
        "Clearances": a["clearances"],
        "Aerial Duels": a["aerials"],
        "Aerials Won": a["aerials_won"],
        "Aerial Win %": round(a["aerials_won"] / a["aerials"] * 100, 1) if a["aerials"] else 0.0,
        "Ball Recoveries": a["recoveries"],
        "Blocked Passes": a["blocked_passes"],
        "Pressures": a["pressures"],
        "Actions In Own Box": a["own_box_actions"],
        "Fouls Committed": a["fouls_committed"],
        "Fouls Won": a["fouls_won"],
        "Cards": a["cards"],
        "Offsides Won": a["offsides_won"],
        "Errors": a["errors"],
    }
    for cat in ZONES:
        record[f"Defensive Actions ({cat})"] = a["zone"][cat]
    records.append(record)

total_df = pd.DataFrame.from_records(records)
total_df.sort_values("Defensive Actions", ascending=False, inplace=True)
total_df.reset_index(drop=True, inplace=True)

# --- per-90 tab ---------------------------------------------------------------
per90_cols = (
    ["Defensive Actions", "Tackles", "Tackles Won", "Dribbled Past", "Interceptions", "Clearances",
     "Aerial Duels", "Aerials Won", "Ball Recoveries", "Blocked Passes", "Pressures", "Actions In Own Box",
     "Fouls Committed", "Fouls Won", "Cards", "Offsides Won", "Errors"]
    + [f"Defensive Actions ({cat})" for cat in ZONES]
)
per90_df = total_df.copy()
factor = per90_df["Minutes"].replace(0, pd.NA) / 90.0
for col in per90_cols:
    per90_df[f"{col}/90"] = (per90_df[col] / factor).round(3)

per90_keep = (
    ["Player", "Team", "Position Group", "Position", "Matches", "Minutes"]
    + [f"{c}/90" for c in per90_cols]
    + ["Tackle Win %", "Aerial Win %"]
)
per90_df = per90_df[per90_keep].fillna(0.0)
per90_df.sort_values("Defensive Actions/90", ascending=False, inplace=True)
per90_df.reset_index(drop=True, inplace=True)

# --- write workbook -------------------------------------------------------
os.makedirs(OUT_DIR, exist_ok=True)
SHEETS = {
    "Per 90": per90_df,
    "Total": total_df,
}
with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
    for sheet_name, df in SHEETS.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# --- formatting -------------------------------------------------------------
from openpyxl import load_workbook

wb = load_workbook(OUT_PATH)
header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E78")
body_font = Font(name="Arial")

for sheet_name in SHEETS:
    ws = wb[sheet_name]
    ws.freeze_panes = "E2"
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(cell.value)), 8)
        for row_cell in ws[col_letter][1:]:
            row_cell.font = body_font
            if isinstance(row_cell.value, (int, float)) and col_idx > 4:
                row_cell.number_format = "0.0" if abs(row_cell.value) >= 10 or row_cell.value == int(row_cell.value) else "0.000"
            max_len = max(max_len, len(str(row_cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)
    ws.auto_filter.ref = ws.dimensions

wb.save(OUT_PATH)
print(f"Wrote {OUT_PATH}: {len(total_df)} players")
