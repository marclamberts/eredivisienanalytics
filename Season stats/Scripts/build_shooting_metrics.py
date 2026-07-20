"""Build Season stats/Shooting Metrics.xlsx from the season's event-derived shot
and minutes data (Danger/all_eredivisie_danger_models.csv, GDA/gda_player_summary.csv,
xT/xt_team_summary.csv, Events/*.json). Two tabs: Per 90 and Total, each with
shots/goals/xG broken out by zone (in box / outside box), situation (open play
/ corner / free kick / penalty), and body part (head / right / left / other),
alongside the overall shooting numbers.

Zone and situation boundaries/detection windows match the conventions already
used in PSV Season Report/Scripts/key_passes_into_box.py and
PSV Season Report/Scripts/season_expansion_pitches.py.
"""
import csv
import glob
import json
import os
from collections import defaultdict, Counter

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUT_DIR = os.path.join(ROOT, "Season stats")
OUT_PATH = os.path.join(OUT_DIR, "Shooting Metrics.xlsx")

BOX_X, BOX_Y_LO, BOX_Y_HI = 83.0, 21.1, 78.9
SHOT_TYPES = {13, 14, 15, 16}
HEAD_QID, RIGHT_QID, LEFT_QID = 15, 20, 72
PENALTY_QID, CORNER_QID, FREEKICK_QID = 9, 6, 5
LINEUP_TYPE_ID = 34
POS_MAP = {1: "GK", 2: "DEF", 3: "MID", 4: "FWD"}


def rows(path):
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def qmap(e):
    return {q.get("qualifierId"): q.get("value") for q in e.get("qualifier", [])}


def in_box(x, y):
    return x >= BOX_X and BOX_Y_LO <= y <= BOX_Y_HI


def build_shot_meta():
    """(match_file, event_id) -> {"situation": ..., "body_part": ...} for every
    shot, replaying each match's timeline to detect corner/free-kick sequences
    exactly as season_expansion_pitches.py does."""
    meta = {}
    for path in sorted(glob.glob(os.path.join(ROOT, "Events", "*.json"))):
        fn = os.path.basename(path)
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        events = [e for e in data.get("event", []) if e.get("periodId") in (1, 2)]
        events.sort(key=lambda e: (e["periodId"], e.get("timeMin", 0), e.get("timeSec", 0), e.get("eventId", 0)))

        last_corner_t, last_freekick_t = {}, {}
        for e in events:
            tid = e.get("typeId")
            cid = e.get("contestantId")
            q = qmap(e)
            t_now = e.get("timeMin", 0) + e.get("timeSec", 0) / 60.0

            if tid == 1:
                if CORNER_QID in q:
                    last_corner_t[cid] = t_now
                if FREEKICK_QID in q:
                    last_freekick_t[cid] = t_now

            if tid in SHOT_TYPES:
                situation = "Open Play"
                if PENALTY_QID in q:
                    situation = "Penalty"
                elif cid in last_corner_t and 0 <= t_now - last_corner_t[cid] <= 0.75:
                    situation = "Corner"
                elif cid in last_freekick_t and 0 <= t_now - last_freekick_t[cid] <= 0.25:
                    situation = "Free Kick"

                if HEAD_QID in q:
                    body_part = "Head"
                elif RIGHT_QID in q:
                    body_part = "Right Foot"
                elif LEFT_QID in q:
                    body_part = "Left Foot"
                else:
                    body_part = "Other"

                meta[(fn, str(e.get("id")))] = {"situation": situation, "body_part": body_part}
    return meta


def build_positions():
    """player_id -> GK/DEF/MID/FWD, from each match's Team Set Up event
    (typeId 34): qualifier 44 gives each of the 23 squad slots' position code
    (1=GK,2=DEF,3=MID,4=FWD,5=unassigned bench), qualifier 131 gives the
    starting slot (1-11) or 0 for bench, and qualifier 30 lists the matching
    player ids -- same lists/order used in Scripts/formation_analysis.py.
    Only starting appearances carry a real position code, so a player's
    most common code across the season is taken as their position."""
    votes = defaultdict(Counter)
    for path in sorted(glob.glob(os.path.join(ROOT, "Events", "*.json"))):
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        for e in data.get("event", []):
            if e.get("typeId") != LINEUP_TYPE_ID:
                continue
            q = qmap(e)
            pids = (q.get(30) or "").split(", ")
            pos = (q.get(44) or "").split(", ")
            slots = (q.get(131) or "").split(", ")
            if not (len(pids) == len(pos) == len(slots)):
                continue
            for pid, p, s in zip(pids, pos, slots):
                if s.strip() != "0" and p.strip() in ("1", "2", "3", "4"):
                    votes[pid][int(p)] += 1
    return {pid: POS_MAP[c.most_common(1)[0][0]] for pid, c in votes.items()}


# --- team names -------------------------------------------------------------
team_names = {}
for r in rows(os.path.join(ROOT, "xT", "xt_team_summary.csv")):
    team_names[r["contestant_id"]] = r["team_name"]

# --- minutes / appearances per player (season totals from per-match rows) ---
minutes = defaultdict(lambda: {"name": "", "contestant_ids": Counter(), "minutes": 0.0, "matches": 0})
for r in rows(os.path.join(ROOT, "GDA", "gda_player_summary.csv")):
    pid = r["player_id"]
    m = minutes[pid]
    m["name"] = r["player_name"]
    m["contestant_ids"][r["contestant_id"]] += 1
    m["minutes"] += num(r["minutes"])
    m["matches"] += int(num(r["matches"]))

# --- shot-level shooting metrics ---------------------------------------------
shot_rows = rows(os.path.join(ROOT, "Danger", "all_eredivisie_danger_models.csv"))
shot_meta = build_shot_meta()
positions = build_positions()

ZONES = ["In Box", "Outside Box"]
SITUATIONS = ["Open Play", "Corner", "Free Kick", "Penalty"]
BODY_PARTS = ["Head", "Right Foot", "Left Foot", "Other"]


def zero_split(keys):
    return {k: {"shots": 0, "goals": 0, "xg": 0.0, "danger": 0.0} for k in keys}


agg = defaultdict(lambda: {
    "name": "", "contestant_ids": Counter(),
    "shots": 0, "sot": 0, "post": 0, "blocked": 0,
    "pen_att": 0, "pen_goals": 0, "goals": 0,
    "xg": 0.0, "npxg": 0.0, "psxg": 0.0, "xgot": 0.0, "danger": 0.0,
    "zone": zero_split(ZONES), "situation": zero_split(SITUATIONS), "body_part": zero_split(BODY_PARTS),
})
for r in shot_rows:
    pid = r["player_id"]
    a = agg[pid]
    a["name"] = r["player_name"] or a["name"] or "Unknown"
    a["contestant_ids"][r["contestant_id"]] += 1
    is_pen = r["is_penalty"] == "1"
    is_goal = r["is_goal"] == "1"
    xg = num(r["xg"])

    a["shots"] += 1
    a["sot"] += int(r["is_on_target"] == "1")
    a["post"] += int(r["is_post"] == "1")
    a["blocked"] += int(r["is_outfield_block"] == "1")
    a["goals"] += int(is_goal)
    a["pen_att"] += int(is_pen)
    a["pen_goals"] += int(is_pen and is_goal)
    a["xg"] += xg
    a["npxg"] += 0.0 if is_pen else xg
    a["psxg"] += num(r["psxg"])
    a["xgot"] += num(r["xgot"])
    danger = num(r["danger_score"])
    a["danger"] += danger

    zone = "In Box" if in_box(num(r["x"]), num(r["y"])) else "Outside Box"
    z = a["zone"][zone]
    z["shots"] += 1
    z["goals"] += int(is_goal)
    z["xg"] += xg
    z["danger"] += danger

    meta = shot_meta.get((r["match_file"], r["event_id"]))
    situation = meta["situation"] if meta else ("Penalty" if is_pen else "Open Play")
    s = a["situation"][situation]
    s["shots"] += 1
    s["goals"] += int(is_goal)
    s["xg"] += xg
    s["danger"] += danger

    body_part = meta["body_part"] if meta else "Other"
    b = a["body_part"][body_part]
    b["shots"] += 1
    b["goals"] += int(is_goal)
    b["xg"] += xg
    b["danger"] += danger

# --- player identity (name / team / matches / minutes), shared by every tab -
player_info = {}
for pid, a in agg.items():
    if a["shots"] == 0:
        continue
    m = minutes.get(pid)
    cid_counter = m["contestant_ids"] if m and m["contestant_ids"] else a["contestant_ids"]
    contestant_id = cid_counter.most_common(1)[0][0] if cid_counter else None
    player_info[pid] = {
        "name": (m["name"] if m else "") or a["name"],
        "team": team_names.get(contestant_id, "Unknown"),
        "position": positions.get(pid, "Unknown"),
        "matches": m["matches"] if m else 0,
        "minutes": m["minutes"] if m else 0.0,
    }

# --- assemble one row per shooter -------------------------------------------
records = []
for pid, a in agg.items():
    if a["shots"] == 0:
        continue
    info = player_info[pid]
    name, team, position, matches, mins = info["name"], info["team"], info["position"], info["matches"], info["minutes"]

    goals, xg, npxg, psxg, xgot, danger = a["goals"], a["xg"], a["npxg"], a["psxg"], a["xgot"], a["danger"]
    np_goals = goals - a["pen_goals"]

    record = {
        "Player": name,
        "Team": team,
        "Position": position,
        "Matches": matches,
        "Minutes": round(mins, 1),
        "Shots": a["shots"],
        "SoT": a["sot"],
        "SoT %": round(a["sot"] / a["shots"] * 100, 1) if a["shots"] else 0.0,
        "Goals": goals,
        "Pen Goals": a["pen_goals"],
        "Pen Att": a["pen_att"],
        "xG": round(xg, 3),
        "npxG": round(npxg, 3),
        "PSxG": round(psxg, 3),
        "xGOT": round(xgot, 3),
        "Danger": round(danger, 3),
        "xG/Shot": round(xg / a["shots"], 3) if a["shots"] else 0.0,
        "Danger/Shot": round(danger / a["shots"], 3) if a["shots"] else 0.0,
        "G-xG": round(goals - xg, 3),
        "npG-npxG": round(np_goals - npxg, 3),
        "G-PSxG": round(goals - psxg, 3),
    }
    for split_key, categories in (("zone", ZONES), ("situation", SITUATIONS), ("body_part", BODY_PARTS)):
        for cat in categories:
            c = a[split_key][cat]
            record[f"Shots ({cat})"] = c["shots"]
            record[f"Goals ({cat})"] = c["goals"]
            record[f"xG ({cat})"] = round(c["xg"], 3)
            record[f"Danger ({cat})"] = round(c["danger"], 3)
    records.append(record)

total_df = pd.DataFrame.from_records(records)
total_df.sort_values("xG", ascending=False, inplace=True)
total_df.reset_index(drop=True, inplace=True)

# --- per-90 tab ---------------------------------------------------------------
split_count_cols = [
    f"{metric} ({cat})"
    for categories in (ZONES, SITUATIONS, BODY_PARTS)
    for cat in categories
    for metric in ("Shots", "Goals", "xG", "Danger")
]
per90_cols = ["Shots", "SoT", "Goals", "Pen Goals", "Pen Att", "xG", "npxG", "PSxG", "xGOT", "Danger"] + split_count_cols
per90_df = total_df.copy()
factor = per90_df["Minutes"].replace(0, pd.NA) / 90.0
for col in per90_cols:
    per90_df[f"{col}/90"] = (per90_df[col] / factor).round(3)
per90_df["G-xG/90"] = (per90_df["G-xG"] / factor).round(3)
per90_df["npG-npxG/90"] = (per90_df["npG-npxG"] / factor).round(3)
per90_df["G-PSxG/90"] = (per90_df["G-PSxG"] / factor).round(3)

per90_keep = (
    ["Player", "Team", "Position", "Matches", "Minutes"]
    + [f"{c}/90" for c in per90_cols]
    + ["SoT %", "xG/Shot", "Danger/Shot", "G-xG/90", "npG-npxG/90", "G-PSxG/90"]
)
per90_df = per90_df[per90_keep].fillna(0.0)
per90_df.sort_values("xG/90", ascending=False, inplace=True)
per90_df.reset_index(drop=True, inplace=True)

# --- write workbook -------------------------------------------------------
os.makedirs(OUT_DIR, exist_ok=True)
SHEETS = {
    "Per 90": per90_df,
    "Total": total_df,
}
with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
    for sheet_name, df in SHEETS.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# --- formatting -------------------------------------------------------------
from openpyxl import load_workbook

wb = load_workbook(OUT_PATH)
header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E78")
body_font = Font(name="Arial")

for sheet_name in SHEETS:
    ws = wb[sheet_name]
    ws.freeze_panes = "D2"
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(cell.value)), 8)
        for row_cell in ws[col_letter][1:]:
            row_cell.font = body_font
            if isinstance(row_cell.value, (int, float)) and col_idx > 3:
                row_cell.number_format = "0.0" if abs(row_cell.value) >= 10 or row_cell.value == int(row_cell.value) else "0.000"
            max_len = max(max_len, len(str(row_cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)
    ws.auto_filter.ref = ws.dimensions

wb.save(OUT_PATH)
print(f"Wrote {OUT_PATH}: {len(total_df)} shooters")
