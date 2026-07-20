"""Build Season stats/Playing Time Metrics.xlsx (starts, sub patterns, and
on-pitch vs off-pitch goal impact) from the season's already-computed GDA
player summary and stint log (GDA/gda_player_summary.csv,
GDA/gda_player_stints.csv), plus position (same lineup-qualifier method as
the other Season stats builders). Two tabs: Per 90 and Total.
"""
import csv
import glob
import json
import os
from collections import defaultdict, Counter

import numpy as np
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUT_DIR = os.path.join(ROOT, "Season stats")
OUT_PATH = os.path.join(OUT_DIR, "Playing Time Metrics.xlsx")

LINEUP_TYPE_ID = 34
POS_MAP = {1: "GK", 2: "DEF", 3: "MID", 4: "FWD"}
LABEL_GROUP = {
    "GK": "GK",
    "CB": "DEF", "LB": "DEF", "RB": "DEF", "LCB": "DEF", "RCB": "DEF", "CB3": "DEF", "LWB": "DEF", "RWB": "DEF",
    "CM": "MID", "LCM": "MID", "RCM": "MID", "LM": "MID", "RM": "MID",
    "ST": "FWD", "LST": "FWD", "RST": "FWD", "LW": "FWD", "RW": "FWD",
}


def label_group(label):
    if label in LABEL_GROUP:
        return LABEL_GROUP[label]
    for prefix in ("DEF", "MID", "FWD"):
        if label.startswith(prefix):
            return prefix
    return None


TOUCH_EXCLUDE = {18, 19, 30, 32, 34, 37, 40, 70, 71, 90, 91}  # non-positional event types


def rows(path):
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def qmap(e):
    return {q.get("qualifierId"): q.get("value") for q in e.get("qualifier", []) or []}


def specific_position_labels(n, group):
    """Left-to-right role names for `n` starters ranked by average touch
    width within `group`. Approximate (there is no official per-role tag in
    the data) -- see build_positions() docstring."""
    if group == "DEF":
        table = {1: ["CB"], 2: ["LB", "RB"], 3: ["LCB", "CB3", "RCB"],
                 4: ["LB", "LCB", "RCB", "RB"], 5: ["LWB", "LCB", "CB", "RCB", "RWB"]}
    elif group == "MID":
        table = {1: ["CM"], 2: ["LCM", "RCM"], 3: ["LM", "CM", "RM"],
                 4: ["LM", "LCM", "RCM", "RM"], 5: ["LM", "LCM", "CM", "RCM", "RM"]}
    elif group == "FWD":
        table = {1: ["ST"], 2: ["LST", "RST"], 3: ["LW", "ST", "RW"]}
    else:
        table = {}
    return table.get(n, [f"{group}{i + 1}" for i in range(n)])


def build_positions():
    """player_id -> (position group, specific position). See the identical
    helper in build_shooting_metrics.py / build_passing_metrics.py for the
    full derivation notes (qualifier 44/131/30 on the Team Set Up event,
    specific role approximated from average touch width per match, then
    restricted to the player's own winning position group)."""
    votes_group = defaultdict(Counter)
    votes_specific = defaultdict(Counter)
    for path in sorted(glob.glob(os.path.join(ROOT, "Events", "*.json"))):
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        events = data.get("event", [])

        touch_y = defaultdict(list)
        for e in events:
            pid, tid, y = e.get("playerId"), e.get("typeId"), e.get("y")
            if pid and y is not None and tid not in TOUCH_EXCLUDE:
                touch_y[pid].append(float(y))

        for e in events:
            if e.get("typeId") != LINEUP_TYPE_ID:
                continue
            q = qmap(e)
            pids = (q.get(30) or "").split(", ")
            pos = (q.get(44) or "").split(", ")
            slots = (q.get(131) or "").split(", ")
            if not (len(pids) == len(pos) == len(slots)):
                continue

            by_group = defaultdict(list)
            for pid, p, s in zip(pids, pos, slots):
                if s.strip() == "0" or p.strip() not in ("1", "2", "3", "4"):
                    continue
                p = int(p)
                votes_group[pid][p] += 1
                if p == 1:
                    votes_specific[pid]["GK"] += 1
                    continue
                avg_y = sum(touch_y[pid]) / len(touch_y[pid]) if touch_y.get(pid) else 50.0
                by_group[POS_MAP[p]].append((pid, avg_y))

            for group, players in by_group.items():
                players.sort(key=lambda t: t[1])
                labels = specific_position_labels(len(players), group)
                for (pid, _), label in zip(players, labels):
                    votes_specific[pid][label] += 1

    groups = {pid: POS_MAP[c.most_common(1)[0][0]] for pid, c in votes_group.items()}
    specific = {}
    for pid, c in votes_specific.items():
        grp = groups.get(pid)
        filtered = Counter({label: n for label, n in c.items() if label_group(label) == grp})
        specific[pid] = (filtered or c).most_common(1)[0][0]
    return groups, specific


# --- team names -------------------------------------------------------------
team_names = {}
for r in rows(os.path.join(ROOT, "xT", "xt_team_summary.csv")):
    team_names[r["contestant_id"]] = r["team_name"]

position_groups, positions = build_positions()

# --- appearance pattern: starts / subbed in / subbed out / red cards --------
appearances = defaultdict(lambda: {"starts": 0, "sub_on": 0, "sub_off": 0, "red_cards": 0, "complete": 0})
for r in rows(os.path.join(ROOT, "GDA", "gda_player_stints.csv")):
    pid = r["player_id"]
    ap = appearances[pid]
    started = r["start_reason"] == "starter"
    ap["starts"] += int(started)
    ap["sub_on"] += int(r["start_reason"] == "sub_on")
    ap["sub_off"] += int(r["end_reason"] == "sub_off")
    ap["red_cards"] += int(r["end_reason"] == "red_card")
    ap["complete"] += int(started and r["end_reason"] == "full_time")

# --- GDA player summary (one row per player-per-club -- a mid-season
# transfer produces two rows for the same player_id, so sum across rows
# rather than keying by player_id straight off) ------------------------------
gda = defaultdict(lambda: {
    "name": "", "contestant_ids": Counter(), "matches": 0, "minutes": 0.0, "off_minutes": 0.0,
    "gf_on": 0.0, "ga_on": 0.0, "gf_off": 0.0, "ga_off": 0.0, "rel_on_off_per90_weighted": 0.0,
})
for r in rows(os.path.join(ROOT, "GDA", "gda_player_summary.csv")):
    pid = r["player_id"]
    g = gda[pid]
    g["name"] = r["player_name"]
    g["contestant_ids"][r["contestant_id"]] += 1
    mins = num(r["minutes"])
    g["matches"] += int(num(r["matches"]))
    g["minutes"] += mins
    g["off_minutes"] += num(r["off_minutes"])
    g["gf_on"] += num(r["goals_for_on"])
    g["ga_on"] += num(r["goals_against_on"])
    g["gf_off"] += num(r["goals_for_off"])
    g["ga_off"] += num(r["goals_against_off"])
    g["rel_on_off_per90_weighted"] += num(r["relative_on_off_goal_difference_per90"]) * mins

# --- assemble one row per player ----------------------------------------------
records = []
for pid, g in gda.items():
    contestant_id = g["contestant_ids"].most_common(1)[0][0] if g["contestant_ids"] else None
    team = team_names.get(contestant_id, "Unknown")
    ap = appearances.get(pid, {"starts": 0, "sub_on": 0, "sub_off": 0, "red_cards": 0, "complete": 0})

    mins = g["minutes"]
    off_mins = g["off_minutes"]
    gf_on, ga_on = g["gf_on"], g["ga_on"]
    gf_off, ga_off = g["gf_off"], g["ga_off"]

    records.append({
        "Player": g["name"],
        "Team": team,
        "Position Group": position_groups.get(pid, "Unknown"),
        "Position": positions.get(pid, "Unknown"),
        "Matches": g["matches"],
        "Starts": ap["starts"],
        "Subbed In": ap["sub_on"],
        "Subbed Out": ap["sub_off"],
        "Complete Matches": ap["complete"],
        "Red Cards": ap["red_cards"],
        "Minutes": round(mins, 1),
        "Off Minutes": round(off_mins, 1),
        "Goals For (On Pitch)": gf_on,
        "Goals Against (On Pitch)": ga_on,
        "GD (On Pitch)": round(gf_on - ga_on, 1),
        "Goals For (Off Pitch)": gf_off,
        "Goals Against (Off Pitch)": ga_off,
        "GD (Off Pitch)": round(gf_off - ga_off, 1),
        "Relative On/Off GD /90": round(g["rel_on_off_per90_weighted"] / mins, 3) if mins else 0.0,
    })

total_df = pd.DataFrame.from_records(records)
total_df.sort_values("Minutes", ascending=False, inplace=True)
total_df.reset_index(drop=True, inplace=True)

# --- per-90 tab (on-pitch rates use Minutes, off-pitch rates use Off Minutes) -
per90_df = total_df.copy()
on_factor = per90_df["Minutes"].replace(0, np.nan) / 90.0
off_factor = per90_df["Off Minutes"].replace(0, np.nan) / 90.0
for col in ["Goals For (On Pitch)", "Goals Against (On Pitch)", "GD (On Pitch)"]:
    per90_df[f"{col}/90"] = (per90_df[col] / on_factor).round(3)
for col in ["Goals For (Off Pitch)", "Goals Against (Off Pitch)", "GD (Off Pitch)"]:
    per90_df[f"{col}/90"] = (per90_df[col] / off_factor).round(3)

per90_keep = (
    ["Player", "Team", "Position Group", "Position", "Matches", "Starts", "Subbed In", "Subbed Out",
     "Complete Matches", "Red Cards", "Minutes", "Off Minutes"]
    + [f"{c}/90" for c in ["Goals For (On Pitch)", "Goals Against (On Pitch)", "GD (On Pitch)",
                            "Goals For (Off Pitch)", "Goals Against (Off Pitch)", "GD (Off Pitch)"]]
    + ["Relative On/Off GD /90"]
)
per90_df = per90_df[per90_keep].fillna(0.0)
per90_df.sort_values("Relative On/Off GD /90", ascending=False, inplace=True)
per90_df.reset_index(drop=True, inplace=True)

# --- write workbook -------------------------------------------------------
os.makedirs(OUT_DIR, exist_ok=True)
SHEETS = {
    "Per 90": per90_df,
    "Total": total_df,
}
with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
    for sheet_name, df in SHEETS.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# --- formatting -------------------------------------------------------------
from openpyxl import load_workbook

wb = load_workbook(OUT_PATH)
header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E78")
body_font = Font(name="Arial")

for sheet_name in SHEETS:
    ws = wb[sheet_name]
    ws.freeze_panes = "E2"
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(cell.value)), 8)
        for row_cell in ws[col_letter][1:]:
            row_cell.font = body_font
            if isinstance(row_cell.value, (int, float)) and col_idx > 4:
                row_cell.number_format = "0.0" if abs(row_cell.value) >= 10 or row_cell.value == int(row_cell.value) else "0.000"
            max_len = max(max_len, len(str(row_cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)
    ws.auto_filter.ref = ws.dimensions

wb.save(OUT_PATH)
print(f"Wrote {OUT_PATH}: {len(total_df)} players")
