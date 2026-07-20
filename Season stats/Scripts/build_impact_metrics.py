"""Build Season stats/Impact Metrics.xlsx (GDA, RAPM, VAEP, Goals Added, xT)
from this season's models. Two tabs: Per 90 and Total.

Prerequisite: run build_rapm.py, build_vaep.py, and build_goals_added.py
first (in Season stats/Scripts/) -- each writes a small _*_cache.csv this
script reads. They're kept separate rather than inlined here because each is
a substantial, independently re-runnable model (RAPM: ridge regression over
lineup segments; VAEP: two gradient-boosted classifiers trained on this
season's ~410k actions; Goals Added: a from-scratch EPV/Markov possession-
value grid). See each script's own docstring for full methodology and
documented simplifications versus the published/described originals --
none of these are the original authors' exact models or trained weights
(Decroos et al.'s VAEP weights and ASA's real g+ model aren't public); they
are this session's own implementations of the published/described methods,
validated by cross-checking against GDA, against each other, and against
known players' real positions where possible (see each script's inline
validation notes).

GDA and xT are this repo's own pre-existing, already-validated models
(GDA/gda_player_summary.csv, xT/xt_player_summary.csv).
"""
import csv
import glob
import json
import os
from collections import defaultdict, Counter

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SCRIPTS_DIR = os.path.join(ROOT, "Season stats", "Scripts")
OUT_DIR = os.path.join(ROOT, "Season stats")
OUT_PATH = os.path.join(OUT_DIR, "Impact Metrics.xlsx")

LINEUP_TYPE_ID = 34
POS_MAP = {1: "GK", 2: "DEF", 3: "MID", 4: "FWD"}
LABEL_GROUP = {
    "GK": "GK",
    "CB": "DEF", "LB": "DEF", "RB": "DEF", "LCB": "DEF", "RCB": "DEF", "CB3": "DEF", "LWB": "DEF", "RWB": "DEF",
    "CM": "MID", "LCM": "MID", "RCM": "MID", "LM": "MID", "RM": "MID",
    "ST": "FWD", "LST": "FWD", "RST": "FWD", "LW": "FWD", "RW": "FWD",
}


def label_group(label):
    if label in LABEL_GROUP:
        return LABEL_GROUP[label]
    for prefix in ("DEF", "MID", "FWD"):
        if label.startswith(prefix):
            return prefix
    return None


TOUCH_EXCLUDE = {18, 19, 30, 32, 34, 37, 40, 70, 71, 90, 91}  # non-positional event types


def rows(path):
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def qmap(e):
    return {q.get("qualifierId"): q.get("value") for q in e.get("qualifier", []) or []}


def specific_position_labels(n, group):
    """Left-to-right role names for `n` starters ranked by average touch
    width within `group`. Approximate (there is no official per-role tag in
    the data) -- see build_positions() docstring."""
    if group == "DEF":
        table = {1: ["CB"], 2: ["LB", "RB"], 3: ["LCB", "CB3", "RCB"],
                 4: ["LB", "LCB", "RCB", "RB"], 5: ["LWB", "LCB", "CB", "RCB", "RWB"]}
    elif group == "MID":
        table = {1: ["CM"], 2: ["LCM", "RCM"], 3: ["LM", "CM", "RM"],
                 4: ["LM", "LCM", "RCM", "RM"], 5: ["LM", "LCM", "CM", "RCM", "RM"]}
    elif group == "FWD":
        table = {1: ["ST"], 2: ["LST", "RST"], 3: ["LW", "ST", "RW"]}
    else:
        table = {}
    return table.get(n, [f"{group}{i + 1}" for i in range(n)])


def build_positions():
    """player_id -> (position group, specific position). See the identical
    helper in build_shooting_metrics.py / build_passing_metrics.py for the
    full derivation notes (qualifier 44/131/30 on the Team Set Up event,
    specific role approximated from average touch width per match, then
    restricted to the player's own winning position group)."""
    votes_group = defaultdict(Counter)
    votes_specific = defaultdict(Counter)
    for path in sorted(glob.glob(os.path.join(ROOT, "Events", "*.json"))):
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        events = data.get("event", [])

        touch_y = defaultdict(list)
        for e in events:
            pid, tid, y = e.get("playerId"), e.get("typeId"), e.get("y")
            if pid and y is not None and tid not in TOUCH_EXCLUDE:
                touch_y[pid].append(float(y))

        for e in events:
            if e.get("typeId") != LINEUP_TYPE_ID:
                continue
            q = qmap(e)
            pids = (q.get(30) or "").split(", ")
            pos = (q.get(44) or "").split(", ")
            slots = (q.get(131) or "").split(", ")
            if not (len(pids) == len(pos) == len(slots)):
                continue

            by_group = defaultdict(list)
            for pid, p, s in zip(pids, pos, slots):
                if s.strip() == "0" or p.strip() not in ("1", "2", "3", "4"):
                    continue
                p = int(p)
                votes_group[pid][p] += 1
                if p == 1:
                    votes_specific[pid]["GK"] += 1
                    continue
                avg_y = sum(touch_y[pid]) / len(touch_y[pid]) if touch_y.get(pid) else 50.0
                by_group[POS_MAP[p]].append((pid, avg_y))

            for group, players in by_group.items():
                players.sort(key=lambda t: t[1])
                labels = specific_position_labels(len(players), group)
                for (pid, _), label in zip(players, labels):
                    votes_specific[pid][label] += 1

    groups = {pid: POS_MAP[c.most_common(1)[0][0]] for pid, c in votes_group.items()}
    specific = {}
    for pid, c in votes_specific.items():
        grp = groups.get(pid)
        filtered = Counter({label: n for label, n in c.items() if label_group(label) == grp})
        specific[pid] = (filtered or c).most_common(1)[0][0]
    return groups, specific


def require_cache(name):
    path = os.path.join(SCRIPTS_DIR, name)
    if not os.path.exists(path):
        raise SystemExit(f"Missing {path} -- run the matching build_*.py script in Season stats/Scripts/ first.")
    return pd.read_csv(path)


# --- team names -------------------------------------------------------------
team_names = {}
for r in rows(os.path.join(ROOT, "xT", "xt_team_summary.csv")):
    team_names[r["contestant_id"]] = r["team_name"]

position_groups, positions = build_positions()

# --- GDA / xT (season-aggregated per player, but with one row per
# player-per-club -- a mid-season transfer produces two rows for the same
# player_id, so sum across rows rather than keying a dict by player_id
# straight off (that would silently keep only the last club's spell). -------
gda = defaultdict(lambda: {
    "name": "", "contestant_ids": Counter(), "matches": 0, "minutes": 0.0,
    "goal_difference_added": 0.0, "rel_gda_per90_weighted": 0.0,
})
for r in rows(os.path.join(ROOT, "GDA", "gda_player_summary.csv")):
    pid = r["player_id"]
    g = gda[pid]
    g["name"] = r["player_name"]
    g["contestant_ids"][r["contestant_id"]] += 1
    mins = num(r["minutes"])
    g["matches"] += int(num(r["matches"]))
    g["minutes"] += mins
    g["goal_difference_added"] += num(r["goal_difference_added"])
    g["rel_gda_per90_weighted"] += num(r["relative_goal_difference_added_per90"]) * mins

xt = defaultdict(lambda: {"xT_added": 0.0, "positive_xT_added": 0.0, "actions": 0})
for r in rows(os.path.join(ROOT, "xT", "xt_player_summary.csv")):
    pid = r["player_id"]
    x = xt[pid]
    x["xT_added"] += num(r["xT_added"])
    x["positive_xT_added"] += num(r["positive_xT_added"])
    x["actions"] += int(num(r["actions"]))

# --- RAPM / VAEP / Goals Added (this session's own models -- see docstring) -
rapm_df = require_cache("_rapm_cache.csv").set_index("player_id")
vaep_df = require_cache("_vaep_cache.csv").set_index("player_id")
ga_df = require_cache("_goals_added_cache.csv").set_index("player_id")
GA_CATEGORIES = ["Dribbling", "Passing", "Receiving", "Shooting", "Interrupting", "Fouling", "Goalkeeping"]

# --- assemble one row per player ----------------------------------------------
records = []
for pid, g in gda.items():
    contestant_id = g["contestant_ids"].most_common(1)[0][0] if g["contestant_ids"] else None
    team = team_names.get(contestant_id, "Unknown")
    matches = g["matches"]
    mins = g["minutes"]
    x = xt.get(pid)

    r_row = rapm_df.loc[pid] if pid in rapm_df.index else None
    v_row = vaep_df.loc[pid] if pid in vaep_df.index else None
    ga_row = ga_df.loc[pid] if pid in ga_df.index else None

    record = {
        "Player": g["name"],
        "Team": team,
        "Position Group": position_groups.get(pid, "Unknown"),
        "Position": positions.get(pid, "Unknown"),
        "Matches": matches,
        "Minutes": round(mins, 1),
        "GDA": round(g["goal_difference_added"], 3),
        "GDA Relative /90": round(g["rel_gda_per90_weighted"] / mins, 3) if mins else 0.0,
        "RAPM": round(float(r_row["RAPM"]), 4) if r_row is not None else 0.0,
        "xRAPM": round(float(r_row["xRAPM"]), 4) if r_row is not None else 0.0,
        "VAEP": round(float(v_row["VAEP"]), 3) if v_row is not None else 0.0,
        "VAEP Offensive": round(float(v_row["vaep_offensive"]), 3) if v_row is not None else 0.0,
        "VAEP Defensive": round(float(v_row["vaep_defensive"]), 3) if v_row is not None else 0.0,
        "Goals Added": round(float(ga_row["Goals Added"]), 3) if ga_row is not None else 0.0,
        "xT Added": round(x["xT_added"], 3) if x else 0.0,
        "Positive xT Added": round(x["positive_xT_added"], 3) if x else 0.0,
        "xT per 100 Actions": round(x["xT_added"] / x["actions"] * 100, 3) if x and x["actions"] else 0.0,
        "Actions": x["actions"] if x else 0,
    }
    for cat in GA_CATEGORIES:
        record[f"GA {cat}"] = round(float(ga_row[cat]), 3) if ga_row is not None else 0.0
    records.append(record)

total_df = pd.DataFrame.from_records(records)
total_df.sort_values("GDA", ascending=False, inplace=True)
total_df.reset_index(drop=True, inplace=True)

# --- per-90 tab -----------------------------------------------------------
# RAPM/xRAPM are already per-90 rates straight out of the regression (the
# target they were fit on was goal/xG differential per 90), so they carry
# through unscaled, same as GDA Relative /90 and xT per 100 Actions.
per90_cols = (
    ["GDA", "VAEP", "VAEP Offensive", "VAEP Defensive", "Goals Added", "xT Added", "Positive xT Added", "Actions"]
    + [f"GA {cat}" for cat in GA_CATEGORIES]
)
per90_df = total_df.copy()
factor = per90_df["Minutes"].replace(0, pd.NA) / 90.0
for col in per90_cols:
    per90_df[f"{col}/90"] = (per90_df[col] / factor).round(3)

per90_keep = (
    ["Player", "Team", "Position Group", "Position", "Matches", "Minutes"]
    + [f"{c}/90" for c in per90_cols]
    + ["GDA Relative /90", "RAPM", "xRAPM", "xT per 100 Actions"]
)
per90_df = per90_df[per90_keep].fillna(0.0)
per90_df.sort_values("GDA/90", ascending=False, inplace=True)
per90_df.reset_index(drop=True, inplace=True)

# --- write workbook -------------------------------------------------------
os.makedirs(OUT_DIR, exist_ok=True)
SHEETS = {
    "Per 90": per90_df,
    "Total": total_df,
}
with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
    for sheet_name, df in SHEETS.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# --- formatting -------------------------------------------------------------
from openpyxl import load_workbook
from openpyxl.comments import Comment

METHOD_NOTES = {
    "RAPM": "Ridge regression of goal differential per 90 on which players shared the pitch, over exact lineup "
            "segments (GDA/gda_player_stints.csv). This season's data only -- wide uncertainty, especially for "
            "low-minute players; the ridge penalty shrinks them toward 0. See build_rapm.py.",
    "xRAPM": "Same as RAPM but the regression target is shot xG differential per 90 (Danger CSV) instead of goals "
              "-- much less sparse than raw goals, so more stable read than RAPM. See build_rapm.py.",
    "VAEP": "This session's own implementation of Decroos et al.'s VAEP framework (gradient-boosted classifiers "
            "predicting goal-scored/conceded within the next 10 actions), trained on this season's ~410k actions. "
            "Held-out AUC ~0.70 for both classifiers. Known limitation shared with the published method: vanilla "
            "VAEP undervalues goalkeepers (their actions sit in states with elevated concede-probability by "
            "construction). See build_vaep.py.",
    "Goals Added": "This session's own from-scratch Expected Possession Value (EPV) model over a 16x12 pitch grid, "
                   "inspired by (not identical to) American Soccer Analysis's publicly described g+ approach -- "
                   "their exact model and category weights aren't public. See build_goals_added.py for documented "
                   "simplifications (dribbling = take-ons only, 50/50 passer/receiver split, goalkeeping scored "
                   "separately as PSxG minus goals conceded).",
}

wb = load_workbook(OUT_PATH)
header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E78")
body_font = Font(name="Arial")

for sheet_name in SHEETS:
    ws = wb[sheet_name]
    ws.freeze_panes = "E2"
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        header_text = str(cell.value).split("/")[0].strip()
        if header_text in METHOD_NOTES:
            cell.comment = Comment(METHOD_NOTES[header_text], "Season stats")
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(cell.value)), 8)
        for row_cell in ws[col_letter][1:]:
            row_cell.font = body_font
            if isinstance(row_cell.value, (int, float)) and col_idx > 4:
                row_cell.number_format = "0.0" if abs(row_cell.value) >= 10 or row_cell.value == int(row_cell.value) else "0.000"
            max_len = max(max_len, len(str(row_cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)
    ws.auto_filter.ref = ws.dimensions

wb.save(OUT_PATH)
print(f"Wrote {OUT_PATH}: {len(total_df)} players")
