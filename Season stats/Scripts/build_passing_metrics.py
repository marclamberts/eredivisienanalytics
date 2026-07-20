"""Build Season stats/Passing Metrics.xlsx from the season's event-derived pass
data (Events/*.json), xT model output (xT/xt_action_values.csv), shot xG
(Danger/all_eredivisie_danger_models.csv, for xA), and season minutes
(GDA/gda_player_summary.csv). Two tabs: Per 90 and Total.

Conventions (box/final-third boundaries, qualifier IDs, progressive/switch
heuristics) match those already used across PSV Season Report/Scripts
(key_passes_into_box.py, long_passes_final_third.py, switches.py,
progressive_carries_final_third.py, individual_metric_pitches.py).
"""
import csv
import glob
import json
import os
from collections import defaultdict, Counter

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUT_DIR = os.path.join(ROOT, "Season stats")
OUT_PATH = os.path.join(OUT_DIR, "Passing Metrics.xlsx")

BOX_X, BOX_Y_LO, BOX_Y_HI = 83.0, 21.1, 78.9
DEF_THIRD_END, FINAL_THIRD_START = 100 / 3, 200 / 3
WIDE_THIRD = 100 / 3
SWITCH_MIN_DIST_M = 30

CROSS_QID, LONGBALL_QID = 2, 1
CORNER_QID, FREEKICK_QID, THROWIN_QID = 6, 5, 107
SET_PIECE_QIDS = {CORNER_QID, FREEKICK_QID, THROWIN_QID}
BOX_EXCLUDE_QIDS = {CORNER_QID, FREEKICK_QID}  # key_passes_into_box.py allows throw-ins
HEAD_QID, RIGHT_QID, LEFT_QID = 15, 20, 72
KEY_PASS_SHOT_LINK_QID = 55  # on the shot: local eventId of the assisting pass
LINEUP_TYPE_ID = 34
POS_MAP = {1: "GK", 2: "DEF", 3: "MID", 4: "FWD"}
LABEL_GROUP = {
    "GK": "GK",
    "CB": "DEF", "LB": "DEF", "RB": "DEF", "LCB": "DEF", "RCB": "DEF", "CB3": "DEF", "LWB": "DEF", "RWB": "DEF",
    "CM": "MID", "LCM": "MID", "RCM": "MID", "LM": "MID", "RM": "MID",
    "ST": "FWD", "LST": "FWD", "RST": "FWD", "LW": "FWD", "RW": "FWD",
}


def label_group(label):
    if label in LABEL_GROUP:
        return LABEL_GROUP[label]
    for prefix in ("DEF", "MID", "FWD"):
        if label.startswith(prefix):
            return prefix
    return None


TOUCH_EXCLUDE = {18, 19, 30, 32, 34, 37, 40, 70, 71, 90, 91}  # non-positional event types


def rows(path):
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def qmap(e):
    return {q.get("qualifierId"): q.get("value") for q in e.get("qualifier", []) or []}


def in_box(x, y):
    return x >= BOX_X and BOX_Y_LO <= y <= BOX_Y_HI


def is_progressive(x0, y0, x1, y1):
    dx = x1 - x0
    return dx >= 10 or (x0 < 50 and x1 >= 66.7) or (x0 < 66.7 and x1 >= 83)


def dist_m(x0, y0, x1, y1):
    dx = (x1 - x0) / 100 * 105.0
    dy = (y1 - y0) / 100 * 68.0
    return (dx ** 2 + dy ** 2) ** 0.5


def is_switch(y0, y1):
    left0, right0 = y0 <= WIDE_THIRD, y0 >= (100 - WIDE_THIRD)
    left1, right1 = y1 <= WIDE_THIRD, y1 >= (100 - WIDE_THIRD)
    return (left0 and right1) or (right0 and left1)


def origin_third(x):
    if x < DEF_THIRD_END:
        return "Defensive Third"
    if x < FINAL_THIRD_START:
        return "Middle Third"
    return "Final Third"


# Through/half-space/zone-14/cutback definitions replicate
# PSV Season Report/Scripts/individual_metric_pitches.py's load_events() exactly
# (success-only, no set-piece exclusion there). Central/Wide extend the same
# start-or-end-in-channel pattern the repo already uses for half-space.
def is_wide_start(y):
    return y <= 21 or y >= 79


def is_box_end(ex, ey):
    return ex >= 83 and 21 <= ey <= 79


def is_cross_hybrid(x, y, ex, ey, q):
    return CROSS_QID in q or (x >= 55 and is_wide_start(y) and is_box_end(ex, ey) and abs(ey - y) >= 12)


def is_through(ex, ey, dx):
    return ex >= 66.7 and dx >= 12 and 33 <= ey <= 67


def is_halfspace(y, ey):
    return (21 <= y < 33 or 67 < y <= 79) or (21 <= ey < 33 or 67 < ey <= 79)


def is_central(y, ey):
    return (33 <= y <= 67) or (33 <= ey <= 67)


def is_wide(y, ey):
    return (y <= 21 or y >= 79) or (ey <= 21 or ey >= 79)


def is_zone14(ex, ey):
    return 66.7 <= ex < 83 and 33 <= ey <= 67


def is_cutback(x, y, ex, ey, q):
    return is_cross_hybrid(x, y, ex, ey, q) and x >= 83 and is_wide_start(y) and ex >= 83 and 33 <= ey <= 67


def specific_position_labels(n, group):
    """Left-to-right role names for `n` starters ranked by average touch
    width within `group`. Approximate (there is no official per-role tag in
    the data) -- see build_positions() docstring."""
    if group == "DEF":
        table = {1: ["CB"], 2: ["LB", "RB"], 3: ["LCB", "CB3", "RCB"],
                 4: ["LB", "LCB", "RCB", "RB"], 5: ["LWB", "LCB", "CB", "RCB", "RWB"]}
    elif group == "MID":
        table = {1: ["CM"], 2: ["LCM", "RCM"], 3: ["LM", "CM", "RM"],
                 4: ["LM", "LCM", "RCM", "RM"], 5: ["LM", "LCM", "CM", "RCM", "RM"]}
    elif group == "FWD":
        table = {1: ["ST"], 2: ["LST", "RST"], 3: ["LW", "ST", "RW"]}
    else:
        table = {}
    return table.get(n, [f"{group}{i + 1}" for i in range(n)])


def build_positions():
    """player_id -> (position group, specific position), from each match's
    Team Set Up event (typeId 34): qualifier 44 gives each of the 23 squad
    slots' position code (1=GK,2=DEF,3=MID,4=FWD,5=unassigned bench),
    qualifier 131 gives the starting slot (1-11) or 0 for bench, and
    qualifier 30 lists the matching player ids -- same lists/order used in
    Scripts/formation_analysis.py. Only starting appearances carry a real
    position code, so a player's most common code across the season becomes
    their position group (GK/DEF/MID/FWD).

    The specific role (e.g. LCB, RM, ST) is not tagged anywhere in the feed,
    so it is approximated per match: within each team's starting XI, players
    sharing a group are ranked by their average y (pitch width) across their
    own touches that match (TOUCH_EXCLUDE mirrors the "touches" definition in
    PSV Season Report/Scripts/season_expansion_pitches.py), then labelled
    left-to-right by squad-role count (e.g. 4 defenders -> LB/LCB/RCB/RB, 3
    -> LCB/CB3/RCB, 5 in a back-five -> LWB/LCB/CB/RCB/RWB). A player's most
    common specific label across the season's starts is taken as their
    position."""
    votes_group = defaultdict(Counter)
    votes_specific = defaultdict(Counter)
    for path in sorted(glob.glob(os.path.join(ROOT, "Events", "*.json"))):
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        events = data.get("event", [])

        touch_y = defaultdict(list)
        for e in events:
            pid, tid, y = e.get("playerId"), e.get("typeId"), e.get("y")
            if pid and y is not None and tid not in TOUCH_EXCLUDE:
                touch_y[pid].append(float(y))

        for e in events:
            if e.get("typeId") != LINEUP_TYPE_ID:
                continue
            q = qmap(e)
            pids = (q.get(30) or "").split(", ")
            pos = (q.get(44) or "").split(", ")
            slots = (q.get(131) or "").split(", ")
            if not (len(pids) == len(pos) == len(slots)):
                continue

            by_group = defaultdict(list)
            for pid, p, s in zip(pids, pos, slots):
                if s.strip() == "0" or p.strip() not in ("1", "2", "3", "4"):
                    continue
                p = int(p)
                votes_group[pid][p] += 1
                if p == 1:
                    votes_specific[pid]["GK"] += 1
                    continue
                avg_y = sum(touch_y[pid]) / len(touch_y[pid]) if touch_y.get(pid) else 50.0
                by_group[POS_MAP[p]].append((pid, avg_y))

            for group, players in by_group.items():
                players.sort(key=lambda t: t[1])
                labels = specific_position_labels(len(players), group)
                for (pid, _), label in zip(players, labels):
                    votes_specific[pid][label] += 1

    groups = {pid: POS_MAP[c.most_common(1)[0][0]] for pid, c in votes_group.items()}
    specific = {}
    for pid, c in votes_specific.items():
        grp = groups.get(pid)
        filtered = Counter({label: n for label, n in c.items() if label_group(label) == grp})
        specific[pid] = (filtered or c).most_common(1)[0][0]
    return groups, specific


# --- team names -------------------------------------------------------------
team_names = {}
for r in rows(os.path.join(ROOT, "xT", "xt_team_summary.csv")):
    team_names[r["contestant_id"]] = r["team_name"]

# --- minutes / appearances per player (season totals) -----------------------
minutes = defaultdict(lambda: {"name": "", "contestant_ids": Counter(), "minutes": 0.0, "matches": 0})
for r in rows(os.path.join(ROOT, "GDA", "gda_player_summary.csv")):
    pid = r["player_id"]
    m = minutes[pid]
    m["name"] = r["player_name"]
    m["contestant_ids"][r["contestant_id"]] += 1
    m["minutes"] += num(r["minutes"])
    m["matches"] += int(num(r["matches"]))

# --- shot xG (for xA) and pass xT, keyed by (match_file, global event id) ---
shot_xg = {}
for r in rows(os.path.join(ROOT, "Danger", "all_eredivisie_danger_models.csv")):
    shot_xg[(r["match_file"], r["event_id"])] = num(r["xg"])

pass_xt = {}
for r in rows(os.path.join(ROOT, "xT", "xt_action_values.csv")):
    if r["move_type"] == "pass":
        pass_xt[(r["match_file"], r["event_id"])] = num(r["xT_added"])

position_groups, positions = build_positions()

# --- splits -------------------------------------------------------------
ORIGIN_THIRDS = ["Defensive Third", "Middle Third", "Final Third"]
PASS_TYPES = ["Open Play", "Corner", "Free Kick", "Throw-in"]
BODY_PARTS = ["Head", "Right Foot", "Left Foot", "Other"]


def zero_split(keys):
    return {k: {"passes": 0, "completed": 0, "xt": 0.0} for k in keys}


agg = defaultdict(lambda: {
    "name": "", "contestant_ids": Counter(),
    "passes": 0, "completed": 0, "xt": 0.0,
    "progressive": 0, "final_third": 0, "box": 0,
    "key_passes": 0, "assists": 0, "xa": 0.0,
    "crosses": 0, "crosses_completed": 0,
    "long_balls": 0, "long_balls_completed": 0,
    "switches": 0,
    "through_balls": 0, "half_space": 0, "central": 0, "wide": 0, "zone14": 0, "cutbacks": 0,
    "origin_third": zero_split(ORIGIN_THIRDS), "type": zero_split(PASS_TYPES), "body_part": zero_split(BODY_PARTS),
})

# --- single pass over every match's events -----------------------------------
for path in sorted(glob.glob(os.path.join(ROOT, "Events", "*.json"))):
    fn = os.path.basename(path)
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    events = data.get("event", [])

    # local eventId (on the assisting pass) -> shot xG, via qualifier 55 on the shot
    assist_pass_xa = {}
    for e in events:
        if e.get("typeId") not in (13, 14, 15, 16):
            continue
        q = qmap(e)
        related = q.get(KEY_PASS_SHOT_LINK_QID)
        if related is None:
            continue
        try:
            related_local_id = int(related)
        except (TypeError, ValueError):
            continue
        xg = shot_xg.get((fn, str(e.get("id"))))
        if xg is not None:
            assist_pass_xa[related_local_id] = xg

    for e in events:
        if e.get("typeId") != 1 or not e.get("playerId"):
            continue
        pid = e["playerId"]
        a = agg[pid]
        a["name"] = e.get("playerName") or a["name"] or "Unknown"
        a["contestant_ids"][e.get("contestantId")] += 1

        q = qmap(e)
        x0, y0 = num(e.get("x")), num(e.get("y"))
        try:
            x1 = float(q.get(140, x0))
        except (TypeError, ValueError):
            x1 = x0
        try:
            y1 = float(q.get(141, y0))
        except (TypeError, ValueError):
            y1 = y0
        success = e.get("outcome") == 1
        xt = pass_xt.get((fn, str(e.get("id"))), 0.0)

        a["passes"] += 1
        a["completed"] += int(success)
        a["xt"] += xt

        if CORNER_QID in q:
            ptype = "Corner"
        elif FREEKICK_QID in q:
            ptype = "Free Kick"
        elif THROWIN_QID in q:
            ptype = "Throw-in"
        else:
            ptype = "Open Play"
        t = a["type"][ptype]
        t["passes"] += 1
        t["completed"] += int(success)
        t["xt"] += xt

        o = a["origin_third"][origin_third(x0)]
        o["passes"] += 1
        o["completed"] += int(success)
        o["xt"] += xt

        if HEAD_QID in q:
            bpart = "Head"
        elif RIGHT_QID in q:
            bpart = "Right Foot"
        elif LEFT_QID in q:
            bpart = "Left Foot"
        else:
            bpart = "Other"
        b = a["body_part"][bpart]
        b["passes"] += 1
        b["completed"] += int(success)
        b["xt"] += xt

        is_open_play = ptype == "Open Play"
        if success and is_open_play:
            if is_progressive(x0, y0, x1, y1):
                a["progressive"] += 1
            if x0 < FINAL_THIRD_START <= x1:
                a["final_third"] += 1
            if not (BOX_EXCLUDE_QIDS & set(q.keys())) and in_box(x1, y1):
                a["box"] += 1
            if not (SET_PIECE_QIDS & set(q.keys())) and is_switch(y0, y1) and dist_m(x0, y0, x1, y1) >= SWITCH_MIN_DIST_M:
                a["switches"] += 1

        if success:
            dx = x1 - x0
            if is_through(x1, y1, dx):
                a["through_balls"] += 1
            if is_halfspace(y0, y1):
                a["half_space"] += 1
            if is_central(y0, y1):
                a["central"] += 1
            if is_wide(y0, y1):
                a["wide"] += 1
            if is_zone14(x1, y1):
                a["zone14"] += 1
            if is_cutback(x0, y0, x1, y1, q):
                a["cutbacks"] += 1

        if CROSS_QID in q:
            a["crosses"] += 1
            a["crosses_completed"] += int(success)
        if LONGBALL_QID in q:
            a["long_balls"] += 1
            a["long_balls_completed"] += int(success)

        if e.get("keyPass") == 1 or e.get("assist") == 1:
            is_assist = e.get("assist") == 1
            a["key_passes"] += 1
            a["assists"] += int(is_assist)
            a["xa"] += assist_pass_xa.get(e.get("eventId"), 0.04 if is_assist else 0.0)

# --- player identity (name / team / matches / minutes), shared by every tab -
player_info = {}
for pid, a in agg.items():
    if a["passes"] == 0:
        continue
    m = minutes.get(pid)
    cid_counter = m["contestant_ids"] if m and m["contestant_ids"] else a["contestant_ids"]
    contestant_id = cid_counter.most_common(1)[0][0] if cid_counter else None
    player_info[pid] = {
        "name": (m["name"] if m else "") or a["name"],
        "team": team_names.get(contestant_id, "Unknown"),
        "position_group": position_groups.get(pid, "Unknown"),
        "position": positions.get(pid, "Unknown"),
        "matches": m["matches"] if m else 0,
        "minutes": m["minutes"] if m else 0.0,
    }

# --- assemble one row per passer ---------------------------------------------
records = []
for pid, a in agg.items():
    if a["passes"] == 0:
        continue
    info = player_info[pid]
    name, team, position_group, position, matches, mins = (
        info["name"], info["team"], info["position_group"], info["position"], info["matches"], info["minutes"]
    )

    record = {
        "Player": name,
        "Team": team,
        "Position Group": position_group,
        "Position": position,
        "Matches": matches,
        "Minutes": round(mins, 1),
        "Passes": a["passes"],
        "Passes Completed": a["completed"],
        "Completion %": round(a["completed"] / a["passes"] * 100, 1) if a["passes"] else 0.0,
        "xT Added": round(a["xt"], 3),
        "Progressive Passes": a["progressive"],
        "Passes Final Third": a["final_third"],
        "Passes Into Box": a["box"],
        "Key Passes": a["key_passes"],
        "Assists": a["assists"],
        "xA": round(a["xa"], 3),
        "Crosses": a["crosses"],
        "Crosses Completed": a["crosses_completed"],
        "Cross %": round(a["crosses_completed"] / a["crosses"] * 100, 1) if a["crosses"] else 0.0,
        "Long Balls": a["long_balls"],
        "Long Balls Completed": a["long_balls_completed"],
        "Long Ball %": round(a["long_balls_completed"] / a["long_balls"] * 100, 1) if a["long_balls"] else 0.0,
        "Switches of Play": a["switches"],
        "Through Balls": a["through_balls"],
        "Half-Space Passes": a["half_space"],
        "Central Passes": a["central"],
        "Wide Passes": a["wide"],
        "Zone 14 Passes": a["zone14"],
        "Cutbacks": a["cutbacks"],
    }
    for split_key, categories in (("origin_third", ORIGIN_THIRDS), ("type", PASS_TYPES), ("body_part", BODY_PARTS)):
        for cat in categories:
            c = a[split_key][cat]
            record[f"Passes ({cat})"] = c["passes"]
            record[f"Completed ({cat})"] = c["completed"]
            record[f"Completion % ({cat})"] = round(c["completed"] / c["passes"] * 100, 1) if c["passes"] else 0.0
            record[f"xT Added ({cat})"] = round(c["xt"], 3)
    records.append(record)

total_df = pd.DataFrame.from_records(records)
total_df.sort_values("xT Added", ascending=False, inplace=True)
total_df.reset_index(drop=True, inplace=True)

# --- per-90 tab ---------------------------------------------------------------
split_count_cols = [
    f"{metric} ({cat})"
    for categories in (ORIGIN_THIRDS, PASS_TYPES, BODY_PARTS)
    for cat in categories
    for metric in ("Passes", "Completed", "xT Added")
]
per90_cols = (
    ["Passes", "Passes Completed", "xT Added", "Progressive Passes", "Passes Final Third", "Passes Into Box",
     "Key Passes", "Assists", "xA", "Crosses", "Crosses Completed", "Long Balls", "Long Balls Completed",
     "Switches of Play", "Through Balls", "Half-Space Passes", "Central Passes", "Wide Passes",
     "Zone 14 Passes", "Cutbacks"]
    + split_count_cols
)
per90_df = total_df.copy()
factor = per90_df["Minutes"].replace(0, pd.NA) / 90.0
for col in per90_cols:
    per90_df[f"{col}/90"] = (per90_df[col] / factor).round(3)

pct_cols = ["Completion %", "Cross %", "Long Ball %"] + [f"Completion % ({cat})" for categories in (ORIGIN_THIRDS, PASS_TYPES, BODY_PARTS) for cat in categories]
per90_keep = ["Player", "Team", "Position Group", "Position", "Matches", "Minutes"] + [f"{c}/90" for c in per90_cols] + pct_cols
per90_df = per90_df[per90_keep].fillna(0.0)
per90_df.sort_values("xT Added/90", ascending=False, inplace=True)
per90_df.reset_index(drop=True, inplace=True)

# --- write workbook -------------------------------------------------------
os.makedirs(OUT_DIR, exist_ok=True)
SHEETS = {
    "Per 90": per90_df,
    "Total": total_df,
}
with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
    for sheet_name, df in SHEETS.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# --- formatting -------------------------------------------------------------
from openpyxl import load_workbook

wb = load_workbook(OUT_PATH)
header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E78")
body_font = Font(name="Arial")

for sheet_name in SHEETS:
    ws = wb[sheet_name]
    ws.freeze_panes = "E2"
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(cell.value)), 8)
        for row_cell in ws[col_letter][1:]:
            row_cell.font = body_font
            if isinstance(row_cell.value, (int, float)) and col_idx > 4:
                row_cell.number_format = "0.0" if abs(row_cell.value) >= 10 or row_cell.value == int(row_cell.value) else "0.000"
            max_len = max(max_len, len(str(row_cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)
    ws.auto_filter.ref = ws.dimensions

wb.save(OUT_PATH)
print(f"Wrote {OUT_PATH}: {len(total_df)} passers")
